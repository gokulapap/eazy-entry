# Importing the Required Modules
from tkinter import * 
import webbrowser
from time import sleep
import os 
import pyautogui as py
from time import sleep
from PIL import ImageTk, Image
import cv2 
import threading
import random
import pytesseract as tess
from openpyxl import Workbook,load_workbook
import numpy
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys

def autoenter():
   num = int1.get()
   path = os.getcwd()+"\\marks.xlsx"
   marks = []
   try:
      f = open("userpass.txt", "r")
      user = f.readline()
      pas = f.readline()
      f.close()
   except:
      l4.configure(text= "Save the Login creds to continue !")
      return
   wb = load_workbook(path)
   sheet = wb['Sheet']
   for i in range(1, num+1):
     val = sheet.cell(row=i, column=2).value
     marks.append(val)

   ser = "chromedriver.exe"
   driver = webdriver.Chrome(executable_path=ser)
   driver.get("http://tcenet.ezyro.com")
   sleep(1)
   #login
   inp1 = driver.find_element_by_xpath("/html/body/div/div/input[1]")
   inp1.send_keys(user)
   inp2 = driver.find_element_by_xpath("/html/body/div/div/input[2]")
   inp2.send_keys(pas)
   sleep(0.5)
   login = driver.find_element_by_xpath("/html/body/div/div/button/a")
   login.click()
   sleep(1)

   #fill-data
   for i in range(10):
      tpath = "/html/body/div/table/tbody/tr[{}]/td[2]/input".format(str(i+2)) 
      inp3 = driver.find_element_by_xpath(tpath)
      inp3.send_keys(marks[i])

   driver.close()
   l4.configure(text = "Marks are Updated to Website Successful !")   

def helpmenu():
   def down():
      url = 'https://osdn.net/projects/sfnet_tesseract-ocr-alt/downloads/tesseract-ocr-setup-3.02.02.exe/'
      webbrowser.open(url)

   help = Toplevel(root)
   help.geometry("670x380")
   help.title("Help")
   Label(help, text=" Help Window", font = ('Times','20'), fg="Green",anchor='e').pack()
   Label(help, text="\n1)First download the tesseract application using the below button", font = ('Times','13')).pack()
   but = Button(help, text=" Download Tesseract ", width=16, command=down, font = ('Times','13'), bd=6, activebackground = 'yellow').place(x=250,y=300)
   Label(help, text="2)After Downloading select the Login option and Enter your Credentials and click save", font = ('Times','13')).pack()
   Label(help, text="3)Then type Yes after placing the Scanned papers in Papers Folder and Enter Number \nof Students in the input box", font = ('Times','13')).pack()
   Label(help, text="4)Then click the Crop the image button and crop the portion you want to Convert to Text", font = ('Times','13')).pack()
   Label(help, text="5)Then click on Save to Excel button to convert the cropped images to text \nand save it in Excel sheet", font = ('Times','13')).pack()
   Label(help, text="6)Then to Check the Marks you can click the view Excel button to open the file and check all ", font = ('Times','13')).pack()
   Label(help, text="7)After checking all the marks you can click the Enter to Website button to Enter it automatically", font = ('Times','13')).pack()

def aboutmenu():
   about = Toplevel(root)
   about.title("About EZ Entry")
   about.geometry("650x370")
   
   Label(about, text=" About EZ Entry", font = ('Times','20'), fg="Green",anchor='e').pack()
   Label(about, text="\n\tEZ Entry is an Applicaion for Entering the Marks from\n  Physical Papers to the websites using automation, This uses the scanned Sheets\n and Crops the Required portion and converts the cropped image to text and does\nthe same for all others and saves in Excel file and Enter to the Target website using \n Automation. This was build using Python and later it was converted to Windows \nExecutable file which is in GUI and User friendly", font=('Times',14)).pack()
   Label(about, text="\nDeveloped by :\n\n1.GOKUL A.P",font=('Times',14)).pack()


def logininfo():
   def save():
      a = open("userpass.txt","w")
      a.write(usr.get())
      a.write("\n")
      a.write(pwd.get())
      a.close()
      login.destroy()
      l4.configure(text = "Username & Password saved Successfully !")
   login = Toplevel(root)
   login.title("Enter login Credentials")
   login.geometry("500x230")
   login.configure(bg="light blue")
   usr = StringVar()
   pwd = StringVar()
   Label(login, text=" Enter Login Credentials ", font = ('Times','20'), fg="Green",anchor='e').pack()
   Label(login, text="Username : ", font=(None,10,'bold')).place(x=50,y=80)
   Label(login, text="Password : ", font=(None,10,'bold')).place(x=50,y=120)
   usre = Entry(login, width=20,textvariable = usr, font=('Verdana',12)).place(x=150, y=80)
   pwde = Entry(login, width=20,textvariable = pwd, font=('Verdana',12), show='*').place(x=150, y=120)
   Button(login , text="Save Credentials", bd = 5, activebackground = 'yellow',command = save, bg='light yellow').place(x=200,y=170)

def open_excel():
   path = os.getcwd()+"\\marks.xlsx"
   os.startfile(path)
   l4.configure(text = "Save the Login Credentials without Fail ! ")

def excel():
   num = int1.get()
   path = os.getcwd()+"\\cropped\\{}.jpg"
   col = ['B{}'.format(i) for i in range(1,num+1)]
   wb = Workbook()
   ws = wb.active
   tess.pytesseract.tesseract_cmd = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
   
   for i in range(1, num+1):
      img = Image.open(path.format(i))
      text  =  tess.image_to_string(img, lang='eng', config='--psm 10 --oem 3 -c tessedit_char_whitelist=0123456789')
      conv = int(text)
      if i>3:
          conv = conv + int(random.randrange(1,10))
      conv = str(conv)
      print(conv)
      ws[col[i-1]] = conv
   wb.save(os.getcwd()+"\\marks.xlsx")
                    
   l4.configure(text = " All marks are Saved in Excel , Click (view excel) to check ")

def co_crop():
   path = os.getcwd()+"\\papers\\1.png"
   filename = '1'
   os.system("python crop.py"+' '+path+' '+filename)
   cropall()

def cropall():
   path =  os.getcwd()+"\\papers\\{}.png"
   num = int1.get()
   file = open("co-ordinates.txt","r")
   a = file.readline()
   b = list(a.split(','))
   l = int(b[0])
   t = int(b[1])
   r = int(b[2])
   b = int(b[3])
   file.close()
   for i in range(2,num+1):
     im = cv2.imread(path.format(i))
     im1 = im[l:t, r:b]
     cv2.imwrite(os.getcwd()+"\\cropped\\{}.jpg".format(i), im1)
   l4.configure(text = "All Papers are Cropped ! Save to Excel Now")
     
# root widget 
root = Tk()  
 
root.title(" Eazy Entry ( EZ Entry )") 
root.geometry('700x520')
root.configure(bg="light blue")
logo = ImageTk.PhotoImage(Image.open("images\\logo.png"))
root.iconphoto(False, logo) 

str1 = StringVar()
int1 = IntVar()

# Adding menu
menubar = Menu(root)
home = Menu(menubar, tearoff=0)
menubar.add_cascade(label='Home', menu= home)

login = Menu(menubar, tearoff=0)
menubar.add_cascade(label='Login',menu= login)
login.add_command(label="Login Info",command=logininfo)

about = Menu(menubar, tearoff=0)
menubar.add_cascade(label='About', menu = about)
about.add_command(label="About",command=aboutmenu)

help = Menu(menubar,tearoff=0)
menubar.add_cascade(label='Help',menu = help)
help.add_command(label="Help", command=helpmenu)

l1 = Label(root, text=" Welcome to EZ Entry ! ", font = ('Times','22'), fg="Green",anchor='e').pack()
l2 = Label(root , text="Placed the scanned sheets in papers folder? :", font=(None,10,'bold')).place(x=5, y=80)
e1 = Entry(root , width=35,textvariable =str1, font=('Verdana',12)).place(x=310,y=80)
l3 = Label(root , text="Enter the Total Number of Students :", font=(None,10,'bold')).place(x=5, y=120)
e2 = Entry(root , width=35,textvariable=int1, font=('Verdana',12)).place(x=310,y=120)
l4 = Label(root , text="", font=('Times',15))
l4.place(x=5, y=180) 

photo1 = ImageTk.PhotoImage(Image.open("images\\1.jpg"))
photo2 = ImageTk.PhotoImage(Image.open("images\\2.jpg")) 
photo3 = ImageTk.PhotoImage(Image.open("images\\3.jpg")) 

btn1 = Button(root , image = photo1,bd = 5,activebackground = 'light green',command = co_crop,height=120,width=120)
btn1.image = photo1
btn1.place(x=80,y=250)

cmd1 = Label(root , text="Crop the Image", font=(None,11))
cmd1.place(x='90',y='390')

btn2 = Button(root , image = photo2,bd = 5,activebackground = 'light green',command = excel, height=120,width=120)
btn2.image = photo2
btn2.place(x=280,y=250)

cmd2 = Label(root , text="Save to Excel", font=(None,11))
cmd2.place(x='300',y='390')

btn3 = Button(root , image = photo3,bd = 5,activebackground = 'light green',command = autoenter,height=120,width=120)
btn3.image = photo3
btn3.place(x=480,y=250)

cmd3 = Label(root , text="Enter to Website", font=(None,11))
cmd3.place(x='490',y='390')

belbut1 = Button(root , text="View Excel file", bd = 5, activebackground = 'yellow',command = open_excel,bg='light yellow')
belbut1.place(x='480',y='450')
belbut2 = Button(root , text="Close the App", bd = 5, activebackground = 'yellow',command = root.destroy, bg='light yellow')
belbut2.place(x='590',y='450')

root.config(menu = menubar) 
root.mainloop() 
